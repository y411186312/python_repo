import sys, os
import json
import Lj_types.my_excel as myExcelLib 
import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string



def get_excel_data(jsonName, excelObj):
    jsonContent =""
    jsonData=None
    with open(jsonName, 'r', encoding='utf-8', errors='ignore') as file:
        jsonData = json.load(file)
        
    excelObj.name          = jsonData["fileName"]
    excelObj.tableCount    = len(jsonData["sheet"])
    
    
    for i in range(excelObj.tableCount):
        sheetJsonObj = jsonData["sheet"][i]
        tableObject = myExcelLib.TableClass();
        tableObject.name = sheetJsonObj["name"]
        #hang
        orgStr = sheetJsonObj["hang"]
        tmpStr = orgStr.replace(" ", "")
        dataList = tmpStr.split(",")
        tableObject.rowCount = len(dataList)
        for j in range(tableObject.rowCount):
            tableObject.rowStrList.append(dataList[j])
        
        #lie 
        orgStr = sheetJsonObj["lie"]
        tmpStr = orgStr.replace(" ", "")
        dataList = tmpStr.split(",")
        print("lie", dataList)
        tableObject.columnCount = len(dataList)
        for j in range(tableObject.columnCount):
            tableObject.columnStrList.append(dataList[j])
        excelObj.tableList.append(tableObject)
   
    print("Excel.name:" , excelObj.name)
    for i in range(excelObj.tableCount):
        tableObj = excelObj.tableList[i]
        print("行数据[%d]-------------" % i)
        for j in range(tableObj.rowCount):
            print(tableObj.rowStrList[j])
        print("列数据[%d]-------------" % i)
        for j in range(tableObj.columnCount):
            print(tableObj.columnStrList[j])



def print_usage(pythonName):
    print("--------------------------------------------------")
    print("Uasge:")
    print("> %s output.xlsx" % pythonName)
    print("需要输入你要输出的excel路径路径")
    print("--------------------------------------------------")


def create_excel_file(excelObj, outputFile):
    wb = openpyxl.Workbook()
    activeSheet = []
    
    #1. create table
    for i in range(excelObj.tableCount):
        #activeSheet = None
        sheetInObj = excelObj.tableList[i]
        sheet = wb.create_sheet(index=i, title=sheetInObj.name)
        
        #2 add content in sheet
        #wb.active = sheet                                   # 设置激活表
        #activeSheet.sheet_properties.tabColor = "205EB2"    # 设置活动表颜色
        #anotherSheet.title = "test"                         # 设置anotherSheet的标题
        sheet.tabColor = "205EB2"    # 设置活动表颜色
        # 行
        for j in range(sheetInObj.columnCount):
            columnDirectionCell = sheet.cell(row=1, column=j + 2)
            columnDirectionCell.value = sheetInObj.columnStrList[j]
        
        # 列
        for j in range(sheetInObj.rowCount):
            rowDirectionCell = sheet.cell(row=j + 2, column=1)
            rowDirectionCell.value = sheetInObj.rowStrList[j]
        #break
    wb.save(excelObj.name+".xlsx")
    
def read_excel_lie_list(wb, table_name, lie_name):
    output = []
    sheet = None
    sheets = wb.sheetnames  #获取所有的sheets
    for i in range(len(sheets)):
        if table_name == sheets[i]:
            print("sheet:", sheets[i])
            sheet = wb[sheets[i]]
            break
    
    if sheet != None:
        max_row = sheet.max_row
        max_column = sheet.max_column
        
         
        print ("最大行：", max_row)
        print ("最大列：", max_column)
        # 遍历工作表中的所有列
        for column in sheet.iter_cols(values_only=True):
            find = False
            for j in range(max_row):
                if column[j] == lie_name:
                    find = True
                if find == True:
                    print(column[j])
    

    return None

def read_excel_whole_row(wb, sheet_name, row_num, column_min, column_max):#读取整行 
    sheets = wb.sheetnames  #获取所有的sheets
    
    sheet = None
    """
    for i in range(len(sheets)):
        print("sheets[%d]:%s" % (i, sheets[i]))
        if sheet_name == sheets[i]:
            sheet = sheets[i]
            break
    """
    try:
        sheet = wb[sheet_name]
        print("Get it")
    except:
        print("not find")
    
    if sheet != None:
        max_row = sheet.max_row
        max_column = sheet.max_column
        print("max_row:%d, max_column:%d" % (max_row, max_column))
        #for cell in list(sheet.rows)[row_num]: #打印单列
        for i in range(column_max-column_min): #打印单列
            print("[%d][%d] : %s "% (row_num, i, sheet.cell(row=row_num, column = i+column_min).value))
    
    """
    rows_max = sheet.max_row
    col_max = sheet.max_column
    
    for i in range(col_max):
        data = sheet[row][i]
    """

#return list, no empty, no duplicate
def read_title_from_sheet(sheet, row_start, row_end, col_char):
    outputList = []
    tempList = []
    col_num = column_index_from_string(col_char)
    for i in range(row_start, row_end+1):
        value = sheet.cell(row=i, column=col_num).value
        if value != None:
            tempList.append(sheet.cell(row=i, column=col_num).value)
            
    for i in range(len(tempList)):
        value = tempList[i]
        find = False
        if i > 0:
            for j in range(len(outputList)):
                if (outputList[j] == tempList[i]):
                    find = True
                    break
        if find == False:
            outputList.append(tempList[i])
        
    for i in range(len(outputList)):
        print("outputList[%d]:%s" % (i, outputList[i]))
    return outputList

def get_cell_row_column(sheet, title):
    max_row     = sheet.max_row
    max_column  = sheet.max_column
    row = 0
    column = 0
    find = False
    
    for row in range(1,max_row+1):
        for column in range(1,max_column+1):
            cell = sheet.cell(row=row, column=column)
            print("cell.value[%d][%d]:%s" % (row, column, cell.value))
            if cell.value == title:
                find = True
                break
        if find == True:
            break
    
    return [row, column, find]    

def save_lie_data_into_file(data_list, fileName):
    with open(fileName, "w") as file:
        for i in range(len(data_list)):
            data = str("%.4d:    %s\n" % (i, data_list[i]))
            file.write(data)
        
 
   
def main(argv):
    appStr = ""
    inputFilePath = ""
    excelObject = myExcelLib.ExcelClass();
    print(sys.argv[0])
    row_num = 2
    data_list = []
    if (len(sys.argv) != 2):
        appStr = sys.argv[0][sys.argv[0].rfind("\\") + 1:]
        
        print_usage(appStr)
    else:
        inputFilePath = sys.argv[1]
        print("File is: " , inputFilePath)
        print("row_num is: " , row_num)
        wb = openpyxl.load_workbook(inputFilePath)
        sheets = wb.sheetnames
        
        
        for i in range(len(sheets)):
            print("Table[%d]:%s" % (i, sheets[i]))
        sheets = wb.sheetnames  #获取所有的sheets
            
        #read_excel_whole_row(wb, "投资支付台账-202401", row_num, 1, 26)
        sheet = wb['Sheet2']
        max_row     = sheet.max_row
        max_column  = sheet.max_column
        row, col, find = get_cell_row_column(sheet, "归档凭证编号")
        print("row:",row)
        print("col:",col)
        print("find:",find)
        print ("max_row: ", sheet.max_row)
        print ("max_column: ", sheet.max_column)
        data_list = read_title_from_sheet(sheet, row+1, max_row, 'C')   #row+1, 下一行才是数据，当前行至少title
        save_lie_data_into_file(data_list, "output.txt")
        #read_title_from_sheet(sheet, 2, 86, 'C')
        #read_excel_lie_list(wb, "投资支付台账-202401", "G")
        #contentList = 
        #get_excel_data(g_jsonFileName, excelObject) 
        #create_excel_file(excelObject, outputFilePath)
    
    #Read_excel("./data/input.xlsx")
    #write_excel("./data/output.xlsx")
    pass
    	
if __name__ == "__main__":
    main(sys.argv)	#main(sys.argv[1:]) means point??? skip app name
    pass 