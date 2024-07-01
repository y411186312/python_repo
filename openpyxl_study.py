
#1. open
wb = openpyxl.load_workbook(r"example.xlsx") # wb means workbook

#2. 从工作簿中取得工作表

sheet_names = wb.get_sheet_names()      # 变成了 sheet_names = wb.sheetnames   sheet_names是一个list
sheet3 = wb.get_sheet_by_name('Sheet3') # sheet3 = wb['Sheet3']  sheet3的type是 openpyxl.worksheet.worksheet.Worksheet
                                        # 获取第n张表  sheet1 = wb[sheet_names[0]]
                                        # 获取具体某个单元 cell = sheet1['A1'] #表示获取第一行第A列， 
                                        # 要得到某个cell的值需要用 cell.value 去获得，如 print (sheet1['A1'].value)
                                        # 已经知道某个单元格 cell, cell.row 单元格的行， cell.column 单元格的列
                                        # 获取具体的cell除了用B5这样的形势也可以用数字形式如 sheet1['B5'] = sheet1.cell(row=5,column=2)
                                        # 
#cell 属性 cell= sheet1['B5']
sheet1['B5'] # Cell 提取单元格
sheet1['B5'].value  #单元格的数据类型和内容
sheet1['B5'].row    #单元格所在的行 = 5
sheet1['B5'].column #单元格所在的列 = 2，即B
sheet1['B5'].coordinate #单元格所在的列， 得到B5                                   
type(sheet3)                            #the type of sheet3 is a worksheet
sheet3.title        #view the title of sheet   


#打印单列
sheet1 = wb.get_active_sheet()
for cellObj in list(sheet1.columns)[1]: #打印单列
    print(cellObj.value)
    
    
列字母和数字之间的转化
import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string
get_column_letter(1)
get_column_letter(100)
column_index_from_string('A')
column_index_from_string('AA')    