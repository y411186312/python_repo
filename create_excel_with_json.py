import sys, os
import json
import Lj_types.my_excel as myExcelLib 
import openpyxl


g_jsonFileName="./data/excel.json"


def get_excel_data1(jsonName, excelObj):
    jsonContent =""
    jsonData=None
    with open(jsonName, 'r', encoding='utf-8', errors='ignore') as file:
        jsonData = json.load(file)
        
    excelObj.name          = jsonData["fileName"]
    excelObj.tableCount    = jsonData["tableCount"]
    
    for i in range(excelObj.tableCount):
        sheetStr="sheet"+str(i)
        sheetJson = jsonData[sheetStr]
        print("--------------------------------------")
        tableObject = myExcelLib.TableClass();
        for j in range(len(sheetJson)):
            isContent = sheetJson[j]["type"]
            nameStr   = sheetJson[j]["name"]
            if (isContent == "biao_ming"):
                tableObject.name=nameStr
                
            elif (isContent == "lie_ming"): #列
                tableObject.columnStrList.append(nameStr)
                tableObject.columCount += 1
            elif (isContent == "hang_ming"): #行
                tableObject.rowStrList.append(nameStr)
                tableObject.rowCount += 1
        excelObj.tableList.append(tableObject)
   
    print("Excel.name:" , excelObj.name)
    for i in range(excelObj.tableCount):
        tableObj = excelObj.tableList[i]
        print("行数据[%d]-------------" % i)
        for j in range(tableObj.rowCount):
            print(tableObj.rowStrList[j])
        print("列数据[%d]-------------" % i)
        for j in range(tableObj.columCount):
            print(tableObj.columnStrList[j])

def get_excel_data(jsonName, excelObj):
    jsonContent =""
    jsonData=None
    with open(jsonName, 'r', encoding='utf-8', errors='ignore') as file:
        jsonData = json.load(file)
        
    excelObj.name          = jsonData["fileName"]
    excelObj.tableCount    = len(jsonData["sheet"])
    
    
    for i in range(excelObj.tableCount):
        sheetJsonObj = jsonData["sheet"][i]
        tableObject = myExcelLib.TableClass();
        tableObject.name = sheetJsonObj["name"]
        #hang
        orgStr = sheetJsonObj["hang"]
        tmpStr = orgStr.replace(" ", "")
        dataList = tmpStr.split(",")
        tableObject.rowCount = len(dataList)
        for j in range(tableObject.rowCount):
            tableObject.rowStrList.append(dataList[j])
        
        #lie 
        orgStr = sheetJsonObj["lie"]
        tmpStr = orgStr.replace(" ", "")
        dataList = tmpStr.split(",")
        print("lie", dataList)
        tableObject.columnCount = len(dataList)
        for j in range(tableObject.columnCount):
            tableObject.columnStrList.append(dataList[j])
        excelObj.tableList.append(tableObject)
   
    print("Excel.name:" , excelObj.name)
    for i in range(excelObj.tableCount):
        tableObj = excelObj.tableList[i]
        print("行数据[%d]-------------" % i)
        for j in range(tableObj.rowCount):
            print(tableObj.rowStrList[j])
        print("列数据[%d]-------------" % i)
        for j in range(tableObj.columnCount):
            print(tableObj.columnStrList[j])



def print_usage(pythonName):
    print("--------------------------------------------------")
    print("Uasge:")
    print("> %s output.xlsx" % pythonName)
    print("需要输入你要输出的excel路径路径")
    print("--------------------------------------------------")


def create_excel_file(excelObj, outputFile):
    wb = openpyxl.Workbook()
    activeSheet = []
    
    #1. create table
    for i in range(excelObj.tableCount):
        #activeSheet = None
        sheetInObj = excelObj.tableList[i]
        sheet = wb.create_sheet(index=i, title=sheetInObj.name)
        
        #2 add content in sheet
        #wb.active = sheet                                   # 设置激活表
        #activeSheet.sheet_properties.tabColor = "205EB2"    # 设置活动表颜色
        #anotherSheet.title = "test"                         # 设置anotherSheet的标题
        sheet.tabColor = "205EB2"    # 设置活动表颜色
        # 行
        for j in range(sheetInObj.columnCount):
            columnDirectionCell = sheet.cell(row=1, column=j + 2)
            columnDirectionCell.value = sheetInObj.columnStrList[j]
        
        # 列
        for j in range(sheetInObj.rowCount):
            rowDirectionCell = sheet.cell(row=j + 2, column=1)
            rowDirectionCell.value = sheetInObj.rowStrList[j]
        #break
    wb.save(excelObj.name+".xlsx")
 
def read_excelRowContent(FileName, rowList):
    wb = openpyxl.load_workbook(FileName)
    
    
    
 
def main(argv):
    appStr = ""
    outputFilePath = ""
    excelObject = myExcelLib.ExcelClass();
    print(sys.argv[0])
    if (len(sys.argv) != 2):
        appStr = sys.argv[0][sys.argv[0].rfind("\\") + 1:]
        print_usage(appStr)
    else:
        outputFilePath = sys.argv[1]
        get_excel_data(g_jsonFileName, excelObject) 
        create_excel_file(excelObject, outputFilePath)
    
    #Read_excel("./data/input.xlsx")
    #write_excel("./data/output.xlsx")
    pass
    	
if __name__ == "__main__":
    main(sys.argv)	#main(sys.argv[1:]) means point??? skip app name
    pass 